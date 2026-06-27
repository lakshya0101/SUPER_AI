"""CSV read/write helpers for validation workflows."""

import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from utils.logger import get_logger


def read_csv_rows(file_path: Path) -> list[dict[str, str]]:
    """Read CSV rows as dictionaries."""
    with file_path.open(mode="r", encoding="utf-8-sig", newline="") as csv_file:
        sample = csv_file.read(4096)
        csv_file.seek(0)

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t")
        except csv.Error:
            dialect = csv.excel

        return list(csv.DictReader(csv_file, dialect=dialect))


def write_csv_rows(file_path: Path, rows: list[dict[str, Any]]) -> Path | None:
    """Write dictionaries to an Excel workbook."""
    if not rows:
        return None

    fieldnames = list(rows[0].keys())
    report_rows = [_clean_report_row(row) for row in rows]
    return _write_excel_rows(file_path.with_suffix(".xlsx"), report_rows, fieldnames)



def _clean_report_row(row: dict[str, Any]) -> dict[str, str]:
    """Normalize report cells so CSV/XLSX filtering stays clean in Excel."""
    cleaned: dict[str, str] = {}
    for key, value in row.items():
        text = "" if value is None else str(value)
        text = text.replace("\r\n", " | ").replace("\n", " | ").replace("\r", " | ")
        text = re.sub(r"\s+", " ", text).strip()
        cleaned[key] = text
    return cleaned


def _write_excel_rows(
    file_path: Path,
    rows: list[dict[str, Any]],
    fieldnames: list[str],
) -> Path | None:
    """Write dictionaries to an Excel workbook when openpyxl is available."""
    logger = get_logger("csv_handler")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning(
            "Excel output skipped because openpyxl is not installed. "
            "Run: pip install -r requirements.txt"
        )
        return None

    output_path = file_path
    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Validation Results"

        for column_index, fieldname in enumerate(fieldnames, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=fieldname)
            cell.font = Font(bold=True)

        for row_index, row in enumerate(rows, start=2):
            for column_index, fieldname in enumerate(fieldnames, start=1):
                value = row.get(fieldname, "")
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column_index, fieldname in enumerate(fieldnames, start=1):
            values = [fieldname]
            values.extend(str(row.get(fieldname, "")) for row in rows[:100])
            max_length = min(max(len(value) for value in values) + 2, 60)
            if fieldname in {"SuperAI_Response", "Citation_Details", "Matched_Evidence", "Document_Data", "Reason"}:
                max_length = 45
            if fieldname in {"Product_Name", "Question", "Matched_Document"}:
                max_length = min(max(max_length, 22), 55)
            if fieldname in {"Page_Number", "Matched_Page", "Result", "Matched_Citation"}:
                max_length = min(max(max_length, 14), 22)
            worksheet.column_dimensions[get_column_letter(column_index)].width = max(
                max_length,
                12,
            )

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_index].height = 45

        try:
            workbook.save(output_path)
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = file_path.with_name(
                f"{file_path.stem}_{timestamp}{file_path.suffix}"
            )
            workbook.save(output_path)

        logger.info("Excel results saved to %s", output_path)
        return output_path
    finally:
        try:
            workbook.close()
        except UnboundLocalError:
            pass
